import sys
import pandas as pd
from sentence_transformers import SentenceTransformer, util
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

MODEL_NAME = "all-MiniLM-L6-v2"
model = SentenceTransformer(MODEL_NAME)

print("ARGS COUNT:", len(sys.argv))
for i, a in enumerate(sys.argv):
    print(f"argv[{i}] =", a)

def run_semantic_matching(file_path, assertion):
    # Load workbook
    wb = load_workbook(file_path)
    ws = wb["Sheet2"]   # your test cases live here

    # Convert Sheet2 into DataFrame
    data = ws.values
    cols = next(data)
    df = pd.DataFrame(data, columns=cols)

    # Merge all columns into one big text blob
    df["FullText"] = df.apply(lambda row: " ".join([str(x) for x in row if pd.notna(x)]), axis=1)

    # Embeddings
    test_embeddings = model.encode(df["FullText"].tolist(), convert_to_tensor=True)
    query_emb = model.encode(assertion, convert_to_tensor=True)
    cos_scores = util.cos_sim(query_emb, test_embeddings)[0].cpu().numpy()

    # Convert to percentage
    df.insert(0, "Semantic Match %", (cos_scores * 100).round(2))

    # Write back to Sheet2 (overwrite table)
    ws.delete_cols(1)  # remove old column if it exists
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.cell(r_idx, 1, row[0])  # Semantic Match %
    ws.cell(1, 1, "Semantic Match %")

    # Add conditional formatting (red → yellow → green)
    color_rule = ColorScaleRule(start_type="num", start_value=0, start_color="FF0000",
                                mid_type="num", mid_value=50, mid_color="FFFF00",
                                end_type="num", end_value=100, end_color="00FF00")
    ws.conditional_formatting.add(f"A2:A{ws.max_row}", color_rule)

    # Save workbook
    wb.save(file_path)

if __name__ == "__main__":
    file_path = sys.argv[1]
    assertion = sys.argv[2]
    run_semantic_matching(file_path, assertion)
